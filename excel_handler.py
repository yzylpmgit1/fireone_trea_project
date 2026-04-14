#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件处理类

这个类提供了处理Excel文件的基本功能，包括打开、读取、写入和关闭Excel文件。
使用openpyxl库来操作Excel文件，支持.xlsx格式。
"""

import openpyxl
from openpyxl.utils import get_column_letter


class ExcelHandler:
    """
    Excel文件处理类
    
    这个类用于处理Excel文件的各种操作，包括：
    - 打开Excel文件
    - 读取工作表数据
    - 写入数据到工作表
    - 保存和关闭Excel文件
    """
    
    def __init__(self, file_path=None):
        """
        初始化ExcelHandler对象
        
        参数:
            file_path (str, optional): Excel文件的路径. 默认值为None
        """
        self.file_path = file_path  # 存储文件路径
        self.workbook = None  # 工作簿对象
        self.active_sheet = None  # 当前活动工作表
    
    def open_file(self, file_path=None):
        """
        打开Excel文件
        
        参数:
            file_path (str, optional): Excel文件的路径. 如果不提供，则使用初始化时的路径
        
        返回:
            bool: 如果文件成功打开，返回True；否则返回False
        """
        # 如果提供了新的文件路径，则更新文件路径
        if file_path:
            self.file_path = file_path
        
        try:
            # 加载Excel文件
            self.workbook = openpyxl.load_workbook(self.file_path)
            # 获取活动工作表
            self.active_sheet = self.workbook.active
            print(f"成功打开文件: {self.file_path}")
            return True
        except Exception as e:
            print(f"打开文件失败: {e}")
            return False
    
    def create_file(self, file_path=None):
        """
        创建一个新的Excel文件
        
        参数:
            file_path (str, optional): 新Excel文件的路径. 如果不提供，则使用初始化时的路径
        
        返回:
            bool: 如果文件成功创建，返回True；否则返回False
        """
        # 如果提供了新的文件路径，则更新文件路径
        if file_path:
            self.file_path = file_path
        
        try:
            # 创建一个新的工作簿
            self.workbook = openpyxl.Workbook()
            # 获取活动工作表
            self.active_sheet = self.workbook.active
            print(f"成功创建文件: {self.file_path}")
            return True
        except Exception as e:
            print(f"创建文件失败: {e}")
            return False
    
    def get_sheet_names(self):
        """
        获取所有工作表的名称
        
        返回:
            list: 工作表名称的列表
        """
        if not self.workbook:
            print("请先打开或创建文件")
            return []
        
        return self.workbook.sheetnames
    
    def switch_sheet(self, sheet_name):
        """
        切换到指定的工作表
        
        参数:
            sheet_name (str): 工作表的名称
        
        返回:
            bool: 如果切换成功，返回True；否则返回False
        """
        if not self.workbook:
            print("请先打开或创建文件")
            return False
        
        try:
            # 切换到指定的工作表
            self.active_sheet = self.workbook[sheet_name]
            print(f"成功切换到工作表: {sheet_name}")
            return True
        except Exception as e:
            print(f"切换工作表失败: {e}")
            return False
    
    def create_sheet(self, sheet_name):
        """
        创建一个新的工作表
        
        参数:
            sheet_name (str): 新工作表的名称
        
        返回:
            bool: 如果创建成功，返回True；否则返回False
        """
        if not self.workbook:
            print("请先打开或创建文件")
            return False
        
        try:
            # 创建新工作表
            new_sheet = self.workbook.create_sheet(sheet_name)
            # 将新工作表设为活动工作表
            self.active_sheet = new_sheet
            print(f"成功创建工作表: {sheet_name}")
            return True
        except Exception as e:
            print(f"创建工作表失败: {e}")
            return False
    
    def read_cell(self, row, column):
        """
        读取指定单元格的值
        
        参数:
            row (int): 行号（从1开始）
            column (int): 列号（从1开始）
        
        返回:
            单元格的值
        """
        if not self.active_sheet:
            print("请先打开或创建文件")
            return None
        
        try:
            # 读取单元格值
            cell_value = self.active_sheet.cell(row=row, column=column).value
            return cell_value
        except Exception as e:
            print(f"读取单元格失败: {e}")
            return None
    
    def read_range(self, start_row, start_column, end_row, end_column):
        """
        读取指定范围的单元格值
        
        参数:
            start_row (int): 起始行号（从1开始）
            start_column (int): 起始列号（从1开始）
            end_row (int): 结束行号
            end_column (int): 结束列号
        
        返回:
            list: 二维列表，包含指定范围内的所有单元格值
        """
        if not self.active_sheet:
            print("请先打开或创建文件")
            return []
        
        try:
            data = []
            # 遍历指定范围的单元格
            for row in range(start_row, end_row + 1):
                row_data = []
                for column in range(start_column, end_column + 1):
                    cell_value = self.active_sheet.cell(row=row, column=column).value
                    row_data.append(cell_value)
                data.append(row_data)
            return data
        except Exception as e:
            print(f"读取单元格范围失败: {e}")
            return []
    
    def write_cell(self, row, column, value):
        """
        向指定单元格写入值
        
        参数:
            row (int): 行号（从1开始）
            column (int): 列号（从1开始）
            value: 要写入的值
        
        返回:
            bool: 如果写入成功，返回True；否则返回False
        """
        if not self.active_sheet:
            print("请先打开或创建文件")
            return False
        
        try:
            # 写入单元格值
            self.active_sheet.cell(row=row, column=column).value = value
            print(f"成功写入单元格 ({row}, {column}): {value}")
            return True
        except Exception as e:
            print(f"写入单元格失败: {e}")
            return False
    
    def write_range(self, start_row, start_column, data):
        """
        向指定范围写入数据
        
        参数:
            start_row (int): 起始行号（从1开始）
            start_column (int): 起始列号（从1开始）
            data (list): 二维列表，包含要写入的数据
        
        返回:
            bool: 如果写入成功，返回True；否则返回False
        """
        if not self.active_sheet:
            print("请先打开或创建文件")
            return False
        
        try:
            # 遍历数据并写入
            for i, row_data in enumerate(data):
                for j, value in enumerate(row_data):
                    row = start_row + i
                    column = start_column + j
                    self.active_sheet.cell(row=row, column=column).value = value
            print(f"成功写入数据范围，从 ({start_row}, {start_column}) 开始")
            return True
        except Exception as e:
            print(f"写入数据范围失败: {e}")
            return False
    
    def save(self):
        """
        保存Excel文件
        
        返回:
            bool: 如果保存成功，返回True；否则返回False
        """
        if not self.workbook or not self.file_path:
            print("请先打开或创建文件")
            return False
        
        try:
            # 保存文件
            self.workbook.save(self.file_path)
            print(f"成功保存文件: {self.file_path}")
            return True
        except Exception as e:
            print(f"保存文件失败: {e}")
            return False
    
    def close(self):
        """
        关闭Excel文件
        
        返回:
            bool: 如果关闭成功，返回True；否则返回False
        """
        if not self.workbook:
            print("没有打开的文件")
            return False
        
        try:
            # 关闭工作簿
            self.workbook.close()
            # 重置属性
            self.workbook = None
            self.active_sheet = None
            print("成功关闭文件")
            return True
        except Exception as e:
            print(f"关闭文件失败: {e}")
            return False


# 使用示例
if __name__ == "__main__":
    """
    示例代码，演示如何使用ExcelHandler类
    """
    # 创建ExcelHandler实例
    handler = ExcelHandler()
    
    # 创建新的Excel文件
    handler.create_file("example.xlsx")
    
    # 写入一些数据
    handler.write_cell(1, 1, "姓名")
    handler.write_cell(1, 2, "年龄")
    handler.write_cell(1, 3, "成绩")
    
    # 写入多行数据
    data = [
        ["张三", 18, 95],
        ["李四", 19, 88],
        ["王五", 17, 92]
    ]
    handler.write_range(2, 1, data)
    
    # 保存文件
    handler.save()
    
    # 关闭文件
    handler.close()
    
    # 重新打开文件
    handler.open_file("example.xlsx")
    
    # 读取数据
    print("\n读取第一行数据:")
    for i in range(1, 4):
        value = handler.read_cell(1, i)
        print(f"第{i}列: {value}")
    
    # 读取数据范围
    print("\n读取所有数据:")
    all_data = handler.read_range(1, 1, 4, 3)
    for row in all_data:
        print(row)
    
    # 创建新工作表
    handler.create_sheet("Sheet2")
    
    # 写入数据到新工作表
    handler.write_cell(1, 1, "这是Sheet2")
    
    # 保存文件
    handler.save()
    
    # 关闭文件
    handler.close()
    
    print("\n示例完成！")
